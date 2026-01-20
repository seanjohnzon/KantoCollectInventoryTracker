"""Tests for allocation service."""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine

from app.db import create_db_and_tables, get_session_factory
from app.models import Allocation, Transaction
from app.services.allocation import (
    _fuzzy_match_item_name,
    get_allocation_summary,
    import_allocations_from_excel,
)


@pytest.fixture
def engine():
    """Create an in-memory SQLite engine for testing."""
    engine = create_engine("sqlite+pysqlite:///:memory:")
    create_db_and_tables(engine)
    return engine


@pytest.fixture
def session_factory(engine):
    """Create a session factory for testing."""
    return get_session_factory(engine)


@pytest.fixture
def sample_transactions(session_factory):
    """Add sample transactions to database."""
    with session_factory() as session:
        transactions = [
            Transaction(
                order_id="001",
                listing_title="Phantasmal Flames Sleeve",
                quantity_sold=100,
                transaction_amount=Decimal("529.00"),
                transaction_type="ORDER_EARNINGS",
                source_file="test.csv",
            ),
            Transaction(
                order_id="002",
                listing_title="Phantasmal Flames ETB",
                quantity_sold=5,
                transaction_amount=Decimal("265.00"),
                transaction_type="ORDER_EARNINGS",
                source_file="test.csv",
            ),
            Transaction(
                order_id="003",
                listing_title="Mega Evolutions Sleeve",
                quantity_sold=20,
                transaction_amount=Decimal("105.80"),
                transaction_type="ORDER_EARNINGS",
                source_file="test.csv",
            ),
        ]
        for t in transactions:
            session.add(t)
        session.commit()


def test_fuzzy_match_item_name():
    """Test fuzzy matching of item names."""
    inventory_items = [
        "Phantasmal Flames Sleeve",
        "Phantasmal Flames ETB",
        "Mega Evolutions Sleeve",
    ]

    # Exact match
    assert _fuzzy_match_item_name("Phantasmal Flames Sleeve", inventory_items) == "Phantasmal Flames Sleeve"

    # Misspelling correction
    assert _fuzzy_match_item_name("Phantasma Flames Sleeve", inventory_items) == "Phantasmal Flames Sleeve"

    # ETB expansion
    assert _fuzzy_match_item_name("Phantasmal Flames Elite Trainer Box", inventory_items) == "Phantasmal Flames ETB"

    # No match
    assert _fuzzy_match_item_name("Random Item That Doesn't Exist", inventory_items) is None


def test_import_allocations_from_excel_dry_run(session_factory, sample_transactions, tmp_path):
    """Test importing allocations in dry run mode."""
    # Create test Excel file
    excel_path = tmp_path / "test_allocations.xlsx"
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Cihan sheet
    sheet = wb.create_sheet("Cihan")
    sheet.append(["Item Name", "Cost", "Count", "Total Cost"])
    sheet.append(["Phantasma Flames Sleeve", 5.29, 50, "=PRODUCT(B2:C2)"])
    sheet.append(["Mega Evolutions Sleeve", 5.29, 10, "=PRODUCT(B3:C3)"])
    
    # Create Nima sheet
    sheet = wb.create_sheet("Nima")
    sheet.append(["Item Name", "Cost", "Count", "Total Cost"])
    sheet.append(["Phantasma Flames ETB", 53.00, 2, "=PRODUCT(B2:C2)"])
    
    wb.save(excel_path)
    
    with session_factory() as session:
        result = import_allocations_from_excel(
            session=session,
            excel_path=str(excel_path),
            title_match="custom",
            dry_run=True,
        )
        
        # Check results
        assert len(result["matched"]) == 3
        assert result["total_allocated"] == 62
        assert len(result["unmatched"]) == 0
        
        # Verify database wasn't modified (dry run)
        allocations = session.query(Allocation).all()
        assert len(allocations) == 0


def test_import_allocations_from_excel_real(session_factory, sample_transactions, tmp_path):
    """Test importing allocations with real database writes."""
    # Create test Excel file
    excel_path = tmp_path / "test_allocations.xlsx"
    wb = openpyxl.Workbook()
    
    wb.remove(wb.active)
    
    sheet = wb.create_sheet("Cihan")
    sheet.append(["Item Name", "Cost", "Count", "Total Cost"])
    sheet.append(["Phantasma Flames Sleeve", 5.29, 50, "=PRODUCT(B2:C2)"])
    
    wb.save(excel_path)
    
    with session_factory() as session:
        result = import_allocations_from_excel(
            session=session,
            excel_path=str(excel_path),
            title_match="custom",
            dry_run=False,
        )
        
        # Check results
        assert len(result["matched"]) == 1
        assert result["total_allocated"] == 50
        
        # Verify database was modified
        allocations = session.query(Allocation).all()
        assert len(allocations) == 1
        assert allocations[0].owner == "Cihan"
        assert allocations[0].allocated_quantity == 50
        assert allocations[0].unit_cost == Decimal("5.29")


def test_get_allocation_summary(session_factory, sample_transactions):
    """Test getting allocation summary."""
    with session_factory() as session:
        # Add allocations
        allocations = [
            Allocation(
                normalized_item_name="Phantasmal Flames Sleeve",
                owner="Cihan",
                allocated_quantity=50,
                unit_cost=Decimal("5.29"),
                excel_item_name="Phantasma Flames Sleeve",
            ),
            Allocation(
                normalized_item_name="Phantasmal Flames Sleeve",
                owner="Askar",
                allocated_quantity=30,
                unit_cost=Decimal("5.29"),
                excel_item_name="Phantasma Flames Sleeve",
            ),
            Allocation(
                normalized_item_name="Mega Evolutions Sleeve",
                owner="Nima",
                allocated_quantity=20,
                unit_cost=Decimal("5.29"),
                excel_item_name="Mega Evolutions Sleeve",
            ),
        ]
        for a in allocations:
            session.add(a)
        session.commit()
        
        # Get summary
        summary = get_allocation_summary(session, title_match="custom")
        
        # Check totals
        assert summary["total_inventory"] == 125  # 100 + 5 + 20
        assert summary["total_allocated"] == 100  # 50 + 30 + 20
        assert summary["total_unallocated"] == 25  # 125 - 100
        
        # Check owner totals
        assert summary["owner_totals"]["Cihan"]["count"] == 50
        assert summary["owner_totals"]["Askar"]["count"] == 30
        assert summary["owner_totals"]["Nima"]["count"] == 20
        
        # Check allocated items
        assert len(summary["allocated_items"]) == 2  # 2 unique items allocated
        
        # Check unallocated items
        unallocated_names = [item["item_name"] for item in summary["unallocated_items"]]
        assert "Phantasmal Flames Sleeve" in unallocated_names  # 20 remaining
        assert "Phantasmal Flames ETB" in unallocated_names  # 5 remaining (not allocated)


def test_get_allocation_summary_empty(session_factory, sample_transactions):
    """Test getting allocation summary with no allocations."""
    with session_factory() as session:
        summary = get_allocation_summary(session, title_match="custom")
        
        # Check that all items are unallocated
        assert summary["total_inventory"] == 125
        assert summary["total_allocated"] == 0
        assert summary["total_unallocated"] == 125
        assert len(summary["owner_totals"]) == 0
        assert len(summary["allocated_items"]) == 0
        assert len(summary["unallocated_items"]) == 3  # All items unallocated
