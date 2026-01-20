"""Allocation service for managing inventory ownership and pricing."""

from __future__ import annotations

import openpyxl
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import select, delete

from app.models import Allocation, ProductImage
from app.services.reporting import normalize_title, get_item_counts


def get_unit_cost_for_item(session: Session, normalized_item_name: str) -> Decimal:
    """
    Get the unit cost for an item from the ProductImage table.
    
    Args:
        session: SQLAlchemy session.
        normalized_item_name: Normalized item name.
    
    Returns:
        Decimal: Unit cost, or 0.00 if not found.
    """
    product_image = session.execute(
        select(ProductImage)
        .where(ProductImage.normalized_item_name == normalized_item_name)
    ).scalar_one_or_none()
    
    if product_image and product_image.unit_cost:
        return product_image.unit_cost
    
    return Decimal("0.00")


def _fuzzy_match_item_name(excel_name: str, inventory_items: List[str]) -> str | None:
    """
    Find the best matching inventory item name for an Excel item name.

    Args:
        excel_name (str): Item name from Excel.
        inventory_items (List[str]): List of normalized inventory item names.

    Returns:
        str | None: Best matching inventory item name, or None if no match.
    """
    # Normalize the Excel name
    excel_lower = excel_name.lower().strip()
    
    # Common replacements and normalizations
    replacements = {
        'phantasma': 'phantasmal',
        'venasaur': 'venusaur',
        'pokeball': 'poke ball',
        'khangaskhan': 'kangaskhan',
        'kanghaskan': 'kangaskhan',
        'masqurade': 'masquerade',
        'masqerade': 'masquerade',
        'booster box': 'booster bundle',
        '3xpack': '3 pack',
        'sneasel': 'sneasel',
        'upc': 'ultra premium collection',
        'elite trainer box': 'etb',
    }
    
    normalized_excel = excel_lower
    for old, new in replacements.items():
        normalized_excel = normalized_excel.replace(old, new)
    
    # Remove common noise words
    noise_words = ['ex', 'pokemon', 'the']
    for word in noise_words:
        normalized_excel = normalized_excel.replace(f' {word} ', ' ')
    
    normalized_excel = ' '.join(normalized_excel.split())  # Remove extra spaces
    
    # Try exact match first
    for inv_item in inventory_items:
        if inv_item.lower() == excel_lower or inv_item.lower() == normalized_excel:
            return inv_item
    
    # Extract key identifying words (set names, product types, character names)
    key_sets = {'phantasmal', 'destined', 'prismatic', 'mega', 'surging', 'twilight', 
                'stellar', 'crown', 'paldean', 'black', 'white', 'shrouded', 'unova',
                'charizard', 'venusaur', 'latias', 'lucario', 'kangaskhan', 'diancie',
                'moltres', 'melmetal', 'kyurem', 'hydreigon', 'dragapult', 'armarouge',
                'sneasel', 'cottonee', 'whimsicott', 'flames', 'evolutions', 'bolt',
                'flare', 'fables', 'sparks', 'masquerade', 'fates'}
    
    key_products = {'sleeve', 'blister', 'etb', 'bundle', 'tin', 'collection', 'deck',
                    'chest', 'figure', 'plush', 'booster', 'elite', 'trainer', 'box',
                    'premium', 'ultra', 'battle'}
    
    excel_words = set(normalized_excel.split())
    excel_set_words = excel_words & key_sets
    excel_product_words = excel_words & key_products
    
    # Score each inventory item
    best_match = None
    best_score = 0
    
    for inv_item in inventory_items:
        inv_lower = inv_item.lower()
        inv_words = set(inv_lower.split())
        
        # Calculate match score
        score = 0
        
        # Set name matches are very important
        common_sets = excel_set_words & inv_words
        score += len(common_sets) * 3
        
        # Product type matches
        common_products = excel_product_words & inv_words
        score += len(common_products) * 2
        
        # Other common words
        common_words = excel_words & inv_words
        score += len(common_words)
        
        # Substring containment bonus
        if normalized_excel in inv_lower or inv_lower in normalized_excel:
            score += 2
        
        # Penalty for length mismatch
        if abs(len(excel_words) - len(inv_words)) > 3:
            score -= 1
        
        # Must have at least one set word or product word match
        if (common_sets or common_products) and score > best_score:
            best_score = score
            best_match = inv_item
    
    # Require minimum score to avoid bad matches
    if best_score >= 4:
        return best_match
    
    return None


def import_allocations_from_excel(
    session: Session,
    excel_path: str,
    title_match: str = "custom",
    dry_run: bool = False
) -> Dict[str, any]:
    """
    Import allocations from Excel file.

    Args:
        session (Session): SQLAlchemy session.
        excel_path (str): Path to Excel file.
        title_match (str): Title matching mode for inventory.
        dry_run (bool): If True, don't save to database, just return preview.

    Returns:
        Dict: Summary of import with matched, unmatched, over_allocated, and totals.
    """
    xlsx_path = Path(excel_path).expanduser()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    
    # Get current inventory
    inventory_items = get_item_counts(
        session=session,
        group_by_buyer=False,
        include_non_sales=True,
        title_match=title_match,
    )
    
    inventory_dict = {item["listing_title"]: item["quantity_sold"] for item in inventory_items}
    inventory_names = list(inventory_dict.keys())
    
    # Read Excel file
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    matched = []
    unmatched = []
    over_allocated = []
    
    # Track total allocated per item across all owners
    total_allocated_per_item = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        owner = sheet_name
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item_name, cost, count, total = row[:4]
            
            if not item_name or not count or not isinstance(item_name, str):
                continue
            
            item_name = item_name.strip()
            
            # Try to match to inventory
            matched_name = _fuzzy_match_item_name(item_name, inventory_names)
            
            if matched_name:
                available = inventory_dict.get(matched_name, 0)
                allocated_qty = float(count)
                
                # Get normalized name for unit cost lookup
                from app.services.reporting import normalize_title
                normalized_name = normalize_title(matched_name.lower(), "custom")
                
                # Get unit cost from database (prioritize database over Excel)
                db_unit_cost = get_unit_cost_for_item(session, normalized_name)
                final_unit_cost = float(db_unit_cost) if db_unit_cost > 0 else (float(cost) if cost else 0.0)
                
                # Track cumulative allocation for this item
                if matched_name not in total_allocated_per_item:
                    total_allocated_per_item[matched_name] = 0
                total_allocated_per_item[matched_name] += allocated_qty
                
                # Check if total allocation exceeds available
                if total_allocated_per_item[matched_name] > available:
                    over_allocated.append({
                        'excel_name': item_name,
                        'inventory_name': matched_name,
                        'owner': owner,
                        'allocated_quantity': allocated_qty,
                        'available_quantity': available,
                        'total_allocated': total_allocated_per_item[matched_name],
                        'unit_cost': final_unit_cost,
                    })
                else:
                    matched.append({
                        'excel_name': item_name,
                        'inventory_name': matched_name,
                        'owner': owner,
                        'allocated_quantity': allocated_qty,
                        'unit_cost': final_unit_cost,
                        'available_quantity': available,
                    })
            else:
                unmatched.append({
                    'excel_name': item_name,
                    'owner': owner,
                    'allocated_quantity': float(count),
                    'unit_cost': float(cost) if cost else 0.0,
                })
    
    if not dry_run:
        # Clear existing allocations
        session.execute(delete(Allocation))
        
        # Insert only valid allocations (not over-allocated)
        for item in matched:
            allocation = Allocation(
                normalized_item_name=item['inventory_name'],
                owner=item['owner'],
                allocated_quantity=int(item['allocated_quantity']),
                unit_cost=Decimal(str(item['unit_cost'])),
                excel_item_name=item['excel_name'],
            )
            session.add(allocation)
        
        session.commit()
    
    return {
        'matched': matched,
        'unmatched': unmatched,
        'over_allocated': over_allocated,
        'total_allocated': sum(m['allocated_quantity'] for m in matched),
        'total_unmatched': sum(u['allocated_quantity'] for u in unmatched),
        'total_over_allocated': sum(o['allocated_quantity'] for o in over_allocated),
    }


def get_allocation_summary(session: Session, title_match: str = "custom") -> Dict[str, any]:
    """
    Get summary of allocations vs available inventory.

    Args:
        session (Session): SQLAlchemy session.
        title_match (str): Title matching mode.

    Returns:
        Dict: Summary with allocated, unallocated, and owner breakdowns.
    """
    # Get current inventory
    inventory_items = get_item_counts(
        session=session,
        group_by_buyer=False,
        include_non_sales=True,
        title_match=title_match,
    )
    
    inventory_dict = {item["listing_title"]: item for item in inventory_items}
    
    # Get allocations
    allocations = session.execute(select(Allocation)).scalars().all()
    
    # Group allocations by item
    allocations_by_item = {}
    for alloc in allocations:
        if alloc.normalized_item_name not in allocations_by_item:
            allocations_by_item[alloc.normalized_item_name] = []
        allocations_by_item[alloc.normalized_item_name].append(alloc)
    
    allocated_items = []
    unallocated_items = []
    
    for item_name, item_data in inventory_dict.items():
        total_quantity = item_data["quantity_sold"]
        normalized_name = item_data.get("normalized_name", item_name.lower())
        # Look up allocations by normalized name, not display name
        item_allocations = allocations_by_item.get(normalized_name, [])
        total_allocated = sum(a.allocated_quantity for a in item_allocations)
        remaining = total_quantity - total_allocated
        
        item_summary = {
            'item_name': item_name,
            'normalized_name': normalized_name,
            'total_quantity': total_quantity,
            'total_allocated': total_allocated,
            'remaining': remaining,
            'set_name': item_data.get('set_name'),
            'image_url': item_data.get('image_url'),
            'allocations': [
                {
                    'owner': a.owner,
                    'quantity': float(a.allocated_quantity),
                    'unit_cost': float(a.unit_cost),
                }
                for a in item_allocations
            ]
        }
        
        if total_allocated > 0:
            allocated_items.append(item_summary)
        
        if remaining >= 0:
            unallocated_items.append({
                'item_name': item_name,
                'quantity': remaining,
                'set_name': item_data.get('set_name'),
                'image_url': item_data.get('image_url'),
                'unit_cost': float(item_data.get('unit_cost', 0.00)),
                'normalized_name': normalized_name,
            })
    
    # Calculate owner totals
    owner_totals = {}
    for alloc in allocations:
        if alloc.owner not in owner_totals:
            owner_totals[alloc.owner] = {'count': 0, 'items': 0}
        owner_totals[alloc.owner]['count'] += alloc.allocated_quantity
        owner_totals[alloc.owner]['items'] += 1
    
    return {
        'allocated_items': allocated_items,
        'unallocated_items': unallocated_items,
        'owner_totals': owner_totals,
        'total_inventory': sum(item["quantity_sold"] for item in inventory_items),
        'total_allocated': sum(a.allocated_quantity for a in allocations),
        'total_unallocated': sum(item['quantity'] for item in unallocated_items),
    }
